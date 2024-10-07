from main import *
import os
import subprocess
from difflib import SequenceMatcher
from openpyxl import Workbook, load_workbook

file_path = './evaluates/EMPLOYEE.cbl'
def evaluate(file_path):
    main(file_path)
    filename = os.path.basename(file_path)
    output_python_filename = filename.split('.')[0] + "_python_output.py"
    with open('./converted.py', 'r') as source:
        source_content = source.read()
    with open(os.path.join("./outputs1/", output_python_filename), 'w') as f:
        f.write(source_content)
    # -------------------------------------------------------------------------------------
def runpython(file_path):
    filename = os.path.basename(file_path)
    output_python_filename = filename.split('.')[0] + "_python_output.py"
    run_output_filename = filename.split('.')[0] + "_python_run.txt"
    subprocess.run(["python3", os.path.join("./outputs1/", output_python_filename)], check=True, stdout=subprocess.DEVNULL)
    with open(os.path.join("./outputs1/txt", run_output_filename), "w") as f_out:
        result = subprocess.run(["python3", os.path.join("./outputs1/", output_python_filename)], stdout=subprocess.PIPE)
        output = result.stdout.decode('utf-8').strip()
        f_out.write(output)
    # -------------------------------------------------------------------------------------
def runcobol(file_path):
    filename = os.path.basename(file_path)
    output_st_filename = filename.split('.')[0] + "_cobol_output.txt"
    subprocess.run(["cobc", "-x", "-o", "coboltest", file_path], check=True, stdout=subprocess.DEVNULL)
    with open(os.path.join("./outputs1/cobtxt", output_st_filename), "w") as f_out:
        result = subprocess.run(["./coboltest"], stdout=subprocess.PIPE)
        output = result.stdout.decode('utf-8').strip()
        f_out.write(output)
    os.remove("coboltest")
    # -------------------------------------------------------------------------------------
def calculate_similarity(file_path):
    """Calculate similarity percentage between two contents."""
    filename = os.path.basename(file_path)
    output_python_filename = filename.split('.')[0] + "_python_run.txt"
    output_st_filename = filename.split('.')[0] + "_cobol_output.txt"
    with open(os.path.join("./outputs1/txt", output_python_filename), 'r') as f:
        content1 = f.read()
    with open(os.path.join("./outputs1/cobtxt", output_st_filename), 'r') as f:
        content2 = f.read()
    return SequenceMatcher(None, content1, content2).ratio() * 100
def add_to_excel(filename, similarity):
    excel_path = "./similarity_report.xlsx"
    if not os.path.exists(excel_path):
        workbook = Workbook() 
        sheet = workbook.active
        sheet.title = "Similarity Report"
        sheet.append(["Filename", "Similarity Percentage"])
    else:
        workbook = load_workbook(excel_path)
        sheet = workbook.active
    sheet.append([filename+".cbl", similarity])
    workbook.save(excel_path)

def final(file_path):
    file = os.path.basename(file_path)
    filename = file.split('.')[0]
    evaluate(file_path)
    runpython(file_path)
    runcobol(file_path)
    similarity = calculate_similarity(file_path)
    print("evaluated percentage for the \"",filename, "\" file is",similarity)
    add_to_excel(filename, similarity)


if __name__ == '__main__':
    final(file_path)